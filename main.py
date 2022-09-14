import eel
import shutil
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import FORMULAE


eel.init("web")

@eel.expose
def synthesize_activate():
    #Load Workbooks
    workbook1 = load_workbook(filename="EthanEdwardsTime.xlsx", data_only=True)
    workbook2 = load_workbook(filename="EvanCampbellTime.xlsx", data_only=True)
    sheet1 = workbook1.active
    sheet2 = workbook2.active

    #Manipulate Data

   #Create Workbook and Cells
    finalWorkbook = Workbook()
    finalSheet = finalWorkbook.active

    #Headers
    finalSheet["A1"] = "Staff Member"
    finalSheet["B1"] = "Total Mileage"
    finalSheet["C1"] = "Mileage Value"

    #Names
    finalSheet["A2"] = sheet1["A2"].value
    finalSheet["A3"] = sheet2["A2"].value

    #Sum Miles
    finalSheet["B2"] = sheet1["N5"].value
    finalSheet["B3"] = sheet2["N5"].value

    #Sum Value
    finalSheet["C2"] = sheet1["O5"].value
    finalSheet["C3"] = sheet2["O5"].value
    
    finalWorkbook.save(filename="Time_Summary.xlsx")

    #Move File to Desktop
    src_path = r'C:\Users\Ethan\OneDrive - The University of Alabama\Code\Capstone\Python_Ex4\Time_Summary.xlsx'
    dst_path = r'C:\Users\Ethan\Desktop'
    shutil.move(src_path, dst_path)

    print("Synthesize")
    return "Complete, file on desktop"


# Start the index.html file
eel.start("index.html", size=(500, 500), mode='none')
